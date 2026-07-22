import re
from openpyxl.cell.cell import MergedCell


class StructureDetector:

    """
    Detects the structure of a timetable sheet.
    """

    # 3F1A
    # 2D2B
    # 4V11

    BATCH_PATTERN = re.compile(r"^\d[A-Z]\d[A-Z0-9]?$")

    # UEC502L
    # UTA025T
    # UNC503P

    SUBJECT_PATTERN = re.compile(
        r"[A-Z]{3}\d{3}[LPT]?$"
    )

    # B303
    # E204
    # C210
    # LT103
    # VL301
    # GC-I

    ROOM_PATTERN = re.compile(
        r"^[A-Z]{1,3}\d{2,4}([/-][A-Z0-9]+)?$"
    )

    # AKR
    # HDJ
    # SSK/GTJ

    FACULTY_PATTERN = re.compile(
        r"^[A-Z]{2,5}(?:/[A-Z]{2,5})*$"
    )

    def __init__(self, sheet):

        self.sheet = sheet

    # --------------------------------------------------

    def get_cell_value(self, row, col):
        """
        Returns the displayed value even if the
        requested cell belongs to a merged range.
        """

        cell = self.sheet.cell(row=row, column=col)

        if not isinstance(cell, MergedCell):
            return cell.value

        for merged in self.sheet.merged_cells.ranges:

            if (
                merged.min_row <= row <= merged.max_row and
                merged.min_col <= col <= merged.max_col
            ):

                return self.sheet.cell(
                    merged.min_row,
                    merged.min_col
                ).value

        return None

    # --------------------------------------------------

    def find_batch_row(self):

        for row in range(1, self.sheet.max_row + 1):

            values = []

            for col in range(1, self.sheet.max_column + 1):

                value = self.get_cell_value(row, col)

                if value is None:
                    continue

                values.append(str(value).strip().upper())

            if "DAY" in values and "HOURS" in values:
                return row

        return None

    # --------------------------------------------------

    def get_batches(self):

        batches = {}

        for row in range(1, self.sheet.max_row + 1):

            for col in range(1, self.sheet.max_column + 1):

                # IMPORTANT: use the actual cell, not get_cell_value()
                value = self.sheet.cell(row=row, column=col).value

                if value is None:
                    continue

                value = str(value).strip().upper()

                if self.BATCH_PATTERN.fullmatch(value):

                    batches[value] = {
                        "row": row,
                        "column": col
                    }

        return batches
    # --------------------------------------------------

    def is_subject(self, value):

        if value is None:
            return False

        value = str(value).upper()

        return bool(
            self.SUBJECT_PATTERN.search(value)
        )

    # --------------------------------------------------

    def is_room(self, value):

        if value is None:
            return False

        value = str(value).strip().upper()

        if value == "LAB":
            return False

        return bool(
            self.ROOM_PATTERN.match(value)
        )

    # --------------------------------------------------

    def is_faculty(self, value):

        if value is None:
            return False

        value = str(value).strip().upper()

        return bool(
            self.FACULTY_PATTERN.match(value)
        )

    # --------------------------------------------------

    def is_lab(self, value):

        if value is None:
            return False

        return str(value).strip().upper() == "LAB"

    # --------------------------------------------------

    def parse_subject(self, text):
        """
        Converts

        UEC502L
        ->
        code = UEC502
        type = Lecture
        """

        if text is None:
            return None

        text = str(text)

        m = re.search(
            r"([A-Z]{3}\d{3})([LPT])",
            text
        )

        if not m:
            return None

        return {

            "code": m.group(1),

            "type": {

                "L": "Lecture",

                "P": "Practical",

                "T": "Tutorial"

            }[m.group(2)]

        }
    def find_hours_column(self):

        header = self.find_batch_row()

        if header is None:
            return None

        for col in range(1, self.sheet.max_column + 1):

            value = self.get_cell_value(header, col)

            if value is None:
                continue

            if str(value).strip().upper() == "HOURS":
                return col

        return None